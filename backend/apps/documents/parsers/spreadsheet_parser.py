"""CSV and XLSX spreadsheet parsers."""

import csv

import openpyxl
import pandas as pd

from apps.documents.parsers.base import BaseParser, ParsedDocument


class CSVParser(BaseParser):
    SUPPORTED_TYPES = {"text/csv"}

    def supports(self, mime_type: str) -> bool:
        return mime_type in self.SUPPORTED_TYPES

    def parse(self, file_path: str) -> ParsedDocument:
        rows = []
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(" | ".join(row))

        full_text = "\n".join(rows)
        metadata = {
            "parser": "csv",
            "row_count": len(rows),
        }
        return ParsedDocument(text=full_text, metadata=metadata, page_count=1)


class XLSXParser(BaseParser):
    SUPPORTED_TYPES = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    def supports(self, mime_type: str) -> bool:
        return mime_type in self.SUPPORTED_TYPES

    def parse(self, file_path: str) -> ParsedDocument:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                sheets.append(f"## Sheet: {sheet_name}\n" + "\n".join(rows))

        wb.close()

        full_text = "\n\n".join(sheets)
        metadata = {
            "parser": "xlsx",
            "sheet_count": len(wb.sheetnames),
        }
        return ParsedDocument(text=full_text, metadata=metadata, page_count=len(wb.sheetnames))
